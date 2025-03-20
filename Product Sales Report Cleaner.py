import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# Streamlit UI
st.title("Product Sales Report Cleaner")
st.write("This web app cleans the Product Sales Report for Kitchen Uwin")

# Upload the Excel file
uploaded_file = st.file_uploader("Choose the Excel file (media.xls or media.xlsx)", type=["xls", "xlsx"])

def clean_excel(file):
    """Cleans the uploaded Excel file and returns a processed file in-memory."""

    # Determine the correct Excel engine
    if file.name.endswith(".xls"):
        df = pd.read_excel(file, dtype=str, header=2, engine="xlrd")
        df_full = pd.read_excel(file, dtype=str, header=2, engine="xlrd")
        df_header = pd.read_excel(file, dtype=str, header=None, nrows=1, engine="xlrd")
    else:  # Default to openpyxl for .xlsx files
        df = pd.read_excel(file, dtype=str, header=2, engine="openpyxl")
        df_full = pd.read_excel(file, dtype=str, header=2, engine="openpyxl")
        df_header = pd.read_excel(file, dtype=str, header=None, nrows=1, engine="openpyxl")

    # Keep only the required columns
    required_columns = ["Description", "U O M", "Quantity"]
    df = df[required_columns]
    
    # Convert Quantity column to numeric
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors='coerce')

    # Define the allowed department values
    allowed_departments = {
        "Hot Hors D’Oeuvres", "Sandwich Platters", "Platters", "Hot & Ready ",
        "Food", "Bakery Items and Breakfast Pastries", "Salad Bowls "
    }

    # Filter based on Department column
    if "Department" in df_full.columns:
        df = df[df_full["Department"].isin(allowed_departments)]

    # Create an in-memory Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    # Load workbook to modify formatting
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # Insert the title row at the top
    for col_num, value in enumerate(df_header.iloc[0], start=1):
        ws.cell(row=1, column=col_num).value = value

    # Set border style for the cleaned data (starting from row 2)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply border only to data cells (avoid empty columns)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value:
                cell.border = thin_border

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save to in-memory buffer
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output

if uploaded_file:
    cleaned_file = clean_excel(uploaded_file)

    # Preserve original filename
    original_filename = uploaded_file.name
    new_filename = original_filename  # Keep same name

    # Download button
    st.download_button(
        label="Download Processed File",
        data=cleaned_file,
        file_name=new_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
